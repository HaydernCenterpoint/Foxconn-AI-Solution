import io
from datetime import datetime
from typing import Any, Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


_HEADER_FILL = PatternFill("solid", fgColor="1C64F2")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header(ws, row: int, cols: List[str]):
    for c, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def generate_xlsx(
    title: str,
    summary: str,
    kpis: List[Dict[str, Any]],
    chart_data: List[Dict[str, Any]],
    top_alarms: List[Dict[str, Any]],
) -> bytes:
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = title
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary["A2"] = f"Ngày xuất: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary.append([])
    _write_header(ws_summary, 4, ["Chỉ số", "Giá trị"])
    for i, kpi in enumerate(kpis, start=5):
        ws_summary.cell(row=i, column=1, value=kpi.get("label", ""))
        ws_summary.cell(row=i, column=2, value=kpi.get("value", ""))
    ws_summary.append([])
    ws_summary.append(["Nhận xét", summary])
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 50

    # --- Sheet 2: Hourly Production ---
    if chart_data:
        ws_prod = wb.create_sheet("Sản lượng")
        keys = list(chart_data[0].keys())
        _write_header(ws_prod, 1, [k.upper() for k in keys])
        for r, row_data in enumerate(chart_data, start=2):
            for c, k in enumerate(keys, start=1):
                ws_prod.cell(row=r, column=c, value=row_data.get(k, ""))
        for col in ws_prod.columns:
            ws_prod.column_dimensions[col[0].column_letter].width = 18

    # --- Sheet 3: Alarms ---
    if top_alarms:
        ws_alarms = wb.create_sheet("Cảnh báo")
        cols = ["machineName", "severity", "message", "status", "createdAt"]
        _write_header(ws_alarms, 1, [c.upper() for c in cols])
        for r, alarm in enumerate(top_alarms, start=2):
            for c, k in enumerate(cols, start=1):
                ws_alarms.cell(row=r, column=c, value=str(alarm.get(k, "")))
        for col in ws_alarms.columns:
            ws_alarms.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
