"""Excel asset import template generator + importer.

Generates an Excel template at:
  data/asset_import_template.xlsx

Import assets from Excel:
  python -m app.scripts.import_assets --file data/assets_to_import.xlsx

Excel format:
  Sheet "Assets": name, type, parent_external_id, external_id, manufacturer,
                   model_number, serial_number, location_zone, location_area,
                   metadata_json, tags
  Sheet "Relationships": asset_external_id, related_external_id, relationship_type, description
"""
from __future__ import annotations

import argparse
import asyncio
import json
import logging
import uuid
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.db.database import get_db_session
from app.schemas.asset import AssetCreateRequest, AssetType, RelationshipCreateRequest, RelationshipType
from app.services.asset_service import AssetService

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def generate_template(output_path: str = "data/asset_import_template.xlsx"):
    wb = Workbook()

    # ============================================================
    # Sheet 1: Assets
    # ============================================================
    ws = wb.active
    ws.title = "Assets"

    headers = [
        "name", "type", "parent_external_id", "external_id",
        "manufacturer", "model_number", "serial_number",
        "location_zone", "location_area",
        "metadata_json", "tags",
        "installed_at",
    ]

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Sample rows
    samples = [
        ["MKZ Factory",       "plant",  "",           "MKZ-HQ",       "MKZ Corp",       "",                  "",                    "Zone A", "",                    "{}",                                          "flagship;primary", ""],
        ["LS18 — Assembly Line 18", "line", "MKZ-HQ", "LS18",         "Siemens",       "",                  "",                    "Zone A", "",                    '{"cycle_time":"45s","stations":12}',        "high-volume;primary", ""],
        ["Press-001",          "machine", "LS18",     "PRESS-001",    "Schuler",       "SMP-2500",          "SCH-2018-001",        "Zone A", "LS18-F01",            '{"power_rating":"2500kW","spindle_hours":12450}', "press;critical", ""],
        ["Press-001-Temp",     "sensor", "PRESS-001", "SENS-P001-T1", "Pt100",         "",                  "",                    "Zone A", "",                    '{"sensor_type":"temperature","unit":"°C"}', "temperature;press", ""],
    ]
    for row_idx, sample in enumerate(samples, 2):
        for col_idx, value in enumerate(sample, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Notes sheet
    notes = wb.create_sheet("Notes")
    notes_data = [
        ["ASSET IMPORT TEMPLATE — READ BEFORE IMPORTING"],
        [""],
        ["COLUMN DESCRIPTIONS:"],
        ["name               Required. Display name of the asset."],
        ["type               Required. One of: plant, line, machine, sensor"],
        ["parent_external_id The external_id of the parent asset (empty for plant)."],
        ["external_id        Optional. Legacy code (machineCode, lineCode, etc)."],
        ["manufacturer       Optional. Manufacturer name."],
        ["model_number       Optional. Model/part number."],
        ["serial_number      Optional. Serial number."],
        ["location_zone      Optional. Zone in factory (e.g., Zone A, Zone B)."],
        ["location_area      Optional. Area designation (e.g., LS18-F01)."],
        ["metadata_json      Optional. JSON object with type-specific fields. Example:"],
        ['                    {"power_rating":"2500kW","spindle_hours":12450,"next_maintenance_date":"2026-08-15"}'],
        ["tags               Optional. Semicolon-separated list of tags."],
        ["installed_at       Optional. ISO date string (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)."],
        [""],
        ["VALIDATION RULES:"],
        ["1. Plant assets must have empty parent_external_id"],
        ["2. type must be one of: plant, line, machine, sensor"],
        ["3. parent_external_id must reference an existing asset's external_id"],
        ["4. external_id must be unique (no duplicates)"],
        ["5. metadata_json must be valid JSON or empty"],
        [""],
        ["TIPS:"],
        ["- Fill in at least the first 3 rows completely as examples"],
        ["- Delete sample rows before importing real data"],
        ["- Run in dry-run mode first: python import_assets.py --file data.xlsx --dry-run"],
        ["- Check the Relationships sheet below for relationship imports"],
    ]
    for row_idx, row_data in enumerate(notes_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = notes.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif ":" in value:
                cell.font = Font(bold=True)

    # ============================================================
    # Sheet 2: Relationships
    # ============================================================
    ws_rel = wb.create_sheet("Relationships")
    rel_headers = ["asset_external_id", "related_external_id", "relationship_type", "description"]
    for col, header in enumerate(rel_headers, 1):
        cell = ws_rel.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    rel_samples = [
        ["PRESS-001", "PRESS-002", "upstream",   "Press-002 feeds press-001 staging"],
        ["PRESS-001", "RW-001",    "upstream",   "Press supplies to weld station"],
        ["CNC-001",   "RW-001",    "upstream",   "CNC parts feed welding robot"],
        ["PB-001",    "OVEN-001",  "upstream",   "Paint booth feeds cure oven"],
    ]
    for row_idx, sample in enumerate(rel_samples, 2):
        for col_idx, value in enumerate(sample, 1):
            ws_rel.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 50
    ws.column_dimensions["K"].width = 30
    ws.column_dimensions["L"].width = 22

    ws_rel.column_dimensions["A"].width = 22
    ws_rel.column_dimensions["B"].width = 22
    ws_rel.column_dimensions["C"].width = 20
    ws_rel.column_dimensions["D"].width = 40

    import os
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    logger.info(f"Excel template generated: {output_path}")
    return output_path


def parse_metadata(metadata_str: str) -> Dict[str, Any]:
    if not metadata_str or not metadata_str.strip():
        return {}
    try:
        return json.loads(metadata_str)
    except json.JSONDecodeError as e:
        logger.warning(f"Invalid JSON in metadata: {metadata_str!r} — {e}")
        return {}


def parse_tags(tags_str: str) -> List[str]:
    if not tags_str or not tags_str.strip():
        return []
    return [t.strip() for t in tags_str.split(";") if t.strip()]


async def import_from_excel(file_path: str, dry_run: bool = True):
    logger.info(f"Loading workbook: {file_path}")
    wb = load_workbook(file_path, data_only=True)

    if "Assets" not in wb.sheetnames:
        raise ValueError("Workbook must have an 'Assets' sheet")

    ws = wb["Assets"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1])}

    def get(row, col_name, default=None):
        idx = header_map.get(col_name)
        return rows[row][idx] if idx is not None and idx < len(rows[row]) else default

    created_ids: Dict[str, uuid.UUID] = {}

    for idx, row in enumerate(rows):
        name = get(idx, "name")
        if not name:
            continue

        asset_type = get(idx, "type", "").strip().lower()
        if asset_type not in ["plant", "line", "machine", "sensor"]:
            logger.warning(f"Row {idx+2}: Unknown type '{asset_type}', skipping")
            continue

        external_id = get(idx, "external_id") or None
        parent_external_id = get(idx, "parent_external_id") or None

        parent_id = created_ids.get(parent_external_id) if parent_external_id else None

        metadata = parse_metadata(get(idx, "metadata_json") or "")
        tags = parse_tags(get(idx, "tags") or "")

        installed_str = get(idx, "installed_at")
        installed_at = None
        if installed_str:
            try:
                from dateutil.parser import parse as parse_date
                installed_at = parse_date(str(installed_str))
            except Exception:
                pass

        req = AssetCreateRequest(
            name=str(name),
            type=AssetType(asset_type),
            parent_id=parent_id,
            external_id=external_id,
            manufacturer=get(idx, "manufacturer"),
            model_number=get(idx, "model_number"),
            serial_number=get(idx, "serial_number"),
            location_zone=get(idx, "location_zone"),
            location_area=get(idx, "location_area"),
            metadata=metadata,
            tags=tags,
            installed_at=installed_at,
        )

        action = "[DRY-RUN] Would create" if dry_run else "Creating"
        logger.info(f"  {action}: {asset_type} '{name}' (parent={parent_external_id or 'root'}, external_id={external_id})")

        if not dry_run:
            async with get_db_session() as session:
                service = AssetService(session)
                asset = await service.create_asset(req)
                created_ids[external_id or str(asset.id)] = asset.id

    # ============================================================
    # Import Relationships
    # ============================================================
    rel_created = 0
    if "Relationships" in wb.sheetnames:
        ws_rel = wb["Relationships"]
        rel_rows = list(ws_rel.iter_rows(min_row=2, values_only=True))
        rel_header_map = {cell.value: idx for idx, cell in enumerate(ws_rel[1])}

        def get_rel(row, col_name, default=None):
            idx = rel_header_map.get(col_name)
            return rel_rows[row][idx] if idx is not None and idx < len(rel_rows[row]) else default

        for idx, row in enumerate(rel_rows):
            asset_ext = get_rel(idx, "asset_external_id")
            related_ext = get_rel(idx, "related_external_id")
            rel_type_str = get_rel(idx, "relationship_type", "related")
            description = get_rel(idx, "description") or None

            if not asset_ext or not related_ext:
                continue

            asset_id = created_ids.get(asset_ext)
            related_id = created_ids.get(related_ext)

            if not asset_id or not related_id:
                logger.warning(f"Row {idx+2}: Could not resolve assets for relationship {asset_ext} -> {related_ext}")
                continue

            try:
                rel_type = RelationshipType(rel_type_str.strip().lower())
            except ValueError:
                logger.warning(f"Row {idx+2}: Unknown relationship type '{rel_type_str}'")
                continue

            action = "[DRY-RUN] Would link" if dry_run else "Linking"
            logger.info(f"  {action}: {asset_ext} --[{rel_type_str}]--> {related_ext}")

            if not dry_run:
                async with get_db_session() as session:
                    service = AssetService(session)
                    await service.create_relationship(RelationshipCreateRequest(
                        asset_id=asset_id,
                        related_asset_id=related_id,
                        relationship_type=rel_type,
                        description=description,
                    ))
                rel_created += 1

    mode = "DRY-RUN" if dry_run else "LIVE"
    logger.info(f"[{mode}] Import complete: {len(created_ids)} assets, {rel_created} relationships")

    if not dry_run:
        logger.info("All data committed to database.")
    return created_ids


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Asset Excel Import Tool")
    parser.add_argument("--file", "-f", default="data/asset_import_template.xlsx", help="Excel file to import")
    parser.add_argument("--generate", "-g", action="store_true", help="Generate import template")
    parser.add_argument("--dry-run", "-n", action="store_true", default=True, help="Dry-run mode (default: True)")
    parser.add_argument("--live", action="store_true", help="Run live import (disable dry-run)")
    args = parser.parse_args()

    if args.generate:
        generate_template()
    else:
        if args.live:
            asyncio.run(import_from_excel(args.file, dry_run=False))
        else:
            asyncio.run(import_from_excel(args.file, dry_run=True))
